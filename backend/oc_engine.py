"""OC intake engine — the single source of truth for the TMP -> Ninja Van transform.

Originally ported from a Node prototype in ../oc-engine; that harness went stale and was
DELETED 18 Aug 2026 (it still had the pre-lock TMP layout and the pre-27-Jul per-piece row
shape). Nothing replaces it — this module is authoritative. Recover from git if ever needed.

Pure transform, no DB/HTTP: parse a SwipeRx TMP .xlsx for a chosen service (S1/S2/S3),
group into AWBs + PO lines, MPS-expand per AWB, and emit the Ninja Van upload columns
(with the courier link injected into delivery_instructions). Rules + provenance:
../OC Template/OC_ENGINE_EXTRACTION_NOTES.md and ../OC Template/OC_TEMPLATE_AND_ENGINE_GUIDE.md.

Tokens are assigned by the caller (oc.py) so they persist in the DB and back the /c/<token>
route — this module only builds the link string once a URL is known.
"""
from __future__ import annotations

import csv
import io
import json
import os
from datetime import date

import openpyxl

_CFG_PATH = os.path.join(os.path.dirname(__file__), "oc_config.json")
with open(_CFG_PATH, encoding="utf-8") as _f:
    CFG = json.load(_f)

# NV upload column order (field names = the real template's header row).
FWD_COLS = [
    "requested_tracking_number", "global_shipper_id", "service_type", "reference.merchant_order_number",
    "service_level", "from.name", "from.phone_number", "from.address.address1", "from.address.country",
    "to.name", "to.phone_number", "to.address.address1", "to.address.country", "to.address.kecamatan",
    "to.address.city", "to.address.province", "to.address.postcode", "parcel_job.delivery_instructions",
    "parcel_job.delivery_start_date", "parcel_job.delivery_timeslot.start_time",
    "parcel_job.delivery_timeslot.end_time", "parcel_job.delivery_timeslot.timezone",
    "parcel_job.dimensions.weight", "parcel_job.is_pickup_required", "parcel_job.items.0.item_description",
    "parcel_job.items.0.is_dangerous_good", "b2b.documents_required", "bundle_information.total_quantity",
    "bundle_information.requested_piece_tracking_numbers", "parcel_job.insured_value", "corporate.branch_id",
]
RET_EXTRA_COLS = [
    "parcel_job.pickup_date", "parcel_job.pickup_timeslot.start_time", "parcel_job.pickup_timeslot.end_time",
    "parcel_job.pickup_timeslot.timezone", "parcel_job.pickup_instructions",
]


class OcError(ValueError):
    """Raised for problems that make the whole file unusable (bad service, unreadable xlsx)."""


def services() -> list[dict]:
    """Public service list for the intake UI.

    Carries the read-only shipper identity FR-OC1 requires the UI to show once a
    service is picked: shipper id + name + corporate branch. Col B always holds the
    MASTER shipper; the per-service shipper is what `branch_id` selects.
    """
    out = []
    for code, s in CFG["services"].items():
        # A service can be switched off in config without deleting its layout/engine
        # support (S3 Return-Pickup is parked this way, 27 Jul 2026). Disabled services
        # disappear from the API, so the UI can't offer them and intake rejects them
        # as unknown_service.
        if not s.get("enabled", True):
            continue
        out.append({"code": code, "name": s["name"], "movement": s["movement"],
                    "service_level": s["service_level"], "direction": s["direction"],
                    "branch_id": s["branch_id"], "shipper_id": s["shipper_id"],
                    "shipper_name": s.get("shipper_name", ""),
                    "master_shipper_id": CFG["master_shipper_id"],
                    "master_shipper_name": CFG.get("master_shipper_name", "")})
    return out


def _norm(v) -> str:
    """Normalize an openpyxl cell value to a clean string (ints without .0, no sci-notation)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _fit_instr(free_text: str, suffix: str, url: str, limit: int) -> str:
    """Wrap fixed/free text + link in <updated_addr>…<a href=URL>URL</a>, trimming free text to fit.
    No separator is inserted between text and the anchor — rdo_text entries carry their own
    trailing punctuation/spacing (e.g. "...berikut:" attaches directly, "...Return. " has a
    trailing space)."""
    def build(t: str) -> str:
        head = "".join([x for x in (t, suffix) if x])
        return f'<updated_addr>{head}<a href="{url}">{url}</a></updated_addr>'

    if len(build(free_text)) <= limit:
        return build(free_text)
    t = free_text
    while t and len(build(t[:-1] + "…")) > limit:
        t = t[:-1]
    return build(t[:-1] + "…" if t else "")


def _load_ws(file_bytes: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface any xlsx failure as a clean error
        raise OcError(f"could not read .xlsx: {exc}") from exc
    return wb.worksheets[0]


def _get(ws, col: str, row: int) -> str:
    return _norm(ws[f"{col}{row}"].value)


def parse(file_bytes: bytes, service_code: str) -> dict:
    """Parse a TMP file into AWB entities (no tokens/links yet).

    Returns {service, direction, awbs:[...], errors:[...], warnings:[...]}.
    Each AWB dict carries the normalized fields + po_lines + total collies; the caller
    assigns a token/url and then calls build_outputs().
    """
    svc = CFG["services"].get(service_code)
    if not svc:
        raise OcError(f"unknown service {service_code!r} (expected S1|S2|S3)")
    layout = CFG["source_layouts"][svc["layout"]]
    ws = _load_ws(file_bytes)
    errors: list[dict] = []
    warnings: list[dict] = []
    awbs = _parse_return(ws, layout, errors) if svc["direction"] == "return" \
        else _parse_forward(ws, layout, errors)
    return {"service": service_code, "direction": svc["direction"],
            "awbs": awbs, "errors": errors, "warnings": warnings}


def _parse_forward(ws, layout, errors) -> list[dict]:
    c = layout["cols"]
    cap = CFG["max_collies_per_awb"]
    order: list[str] = []
    groups: dict[str, dict] = {}
    current = None
    for r in range(layout["data_start_row"], ws.max_row + 1):
        key = _get(ws, c["swipe_awb"], r)
        po = _get(ws, c["po"], r)
        koli_raw = _get(ws, c["koli"], r)
        if not (key or po or _get(ws, c["name"], r) or koli_raw):
            continue
        if key:
            current = key
            if key not in groups:
                groups[key] = {
                    "awb_id": key, "merchant_order_number": key,
                    "pharmacy_name": _get(ws, c["name"], r), "phone": _get(ws, c["phone"], r),
                    "address": _get(ws, c["address"], r), "city": _get(ws, c["city"], r),
                    "postcode": _get(ws, c["zip"], r), "weight": _get(ws, c["weight"], r),
                    "po_lines": [], "collies": 0, "is_return": False,
                    "invoice": "", "item_detail": "",
                }
                order.append(key)
        if current is None:
            errors.append({"row": r, "awb": "", "error": "orphan row before any SwipeAWB"})
            continue
        try:
            koli = int(float(koli_raw))
        except ValueError:
            koli = 0
        if not po:
            errors.append({"row": r, "awb": current, "error": "missing PO Number"})
            continue
        if koli <= 0:
            errors.append({"row": r, "awb": current, "error": f'koli must be > 0 (got "{koli_raw}")'})
            continue
        g = groups[current]
        g["po_lines"].append({"po_number": po, "koli": koli})
        g["collies"] += koli

    out = []
    for key in order:
        g = groups[key]
        total = g["collies"]
        if total <= 0:
            errors.append({"row": None, "awb": key, "error": "no valid collies"})
            continue
        if total > cap:
            errors.append({"row": None, "awb": key,
                           "error": f"collie count {total} exceeds cap {cap} — likely a layout mismatch; skipped"})
            continue
        g["collies"] = total
        out.append(g)
    return out


def _parse_return(ws, layout, errors) -> list[dict]:
    c = layout["cols"]
    out = []
    for r in range(layout["data_start_row"], ws.max_row + 1):
        return_awb = _get(ws, c["return_awb"], r)
        po = _get(ws, c["po"], r)
        if not return_awb and not po:
            continue
        row_err = []
        if not return_awb:
            row_err.append("missing return AWB")
        if not _get(ws, c["address"], r):
            row_err.append("missing pharmacy address")
        if row_err:
            errors.append({"row": r, "awb": return_awb, "error": "; ".join(row_err)})
            continue
        detail = " ".join(_get(ws, c["detail"], r).split())
        out.append({
            "awb_id": return_awb, "merchant_order_number": po,
            "pharmacy_name": _get(ws, c["pharmacy_name"], r), "phone": _get(ws, c["phone"], r),
            "address": _get(ws, c["address"], r), "city": _get(ws, c["city"], r),
            "postcode": "", "weight": "1",
            "po_lines": [{"po_number": po, "koli": 1}], "collies": 1, "is_return": True,
            "invoice": _get(ws, c["inv"], r), "item_detail": detail,
        })
    return out


def piece_trids(awb: dict) -> list[str]:
    """Child piece TRIDs inside the AWB's single MPS bundle (guide §2.3).

    Forward: one child per collie, prefixed with the **PO Number** that collie belongs to,
    numbered **continuously across the whole AWB** (not restarting per PO) and zero-padded
    to two digits.

        AWB02U24V: PO1(2 koli), PO2(1), PO3(1)
        -> PO1-01, PO1-02, PO2-03, PO3-04

    Return (S3): the pre-supplied AWBR with a single -01 piece.

    PO-prefixed is deliberate. NV permits custom piece ids, and the children still attach to
    their parent when the order is created from the manual template — confirmed by Baskoro,
    10 and 18 Aug 2026. It is also what makes the per-PO RDO / SP-Manual check work at the
    door, since each piece traces to its own PO.

    ⚠️ This was briefly changed to parent-prefixed children on 10 Aug 2026, on the strength of
    an audit showing 1035/1035 children in three accepted BULK uploads used prefix == col A.
    That inference was wrong: it read a property of the bulk-upload path as a universal rule,
    and the manual-template path this account actually uses accepts PO prefixes. Converter v6
    rolled the same change back. Do not "fix" this again without asking.

    These are the CHILDREN only. The bundle's own tracking number is the SwipeAWB and is
    written to cols A and D by `_upload_row` — never taken from this list.
    """
    if awb["is_return"]:
        return [f'{awb["awb_id"]}-01']
    out: list[str] = []
    n = 0
    for p in awb["po_lines"]:
        for _ in range(max(1, int(p["koli"] or 1))):
            n += 1
            out.append(f'{p["po_number"]}-{n:02d}')
    return out


def _fit_forward(mandated: str, cta: str, url: str, limit: int) -> str:
    """Build col R with the mandated RDO wording protected.

    That wording is compliance text (guide §2.4) and must never be silently trimmed —
    only the call-to-action is discretionary, so it is dropped WHOLE when the budget is
    tight. This stopped being theoretical on 27 Jul 2026: the platform moved the app to
    an org-scoped host, the URL grew 9 chars, and since the URL is counted twice the
    text budget fell from 296 to 278 — less than the 296 the old text+CTA needed.

    If even the bare mandated text cannot fit, the result is ellipsised AND reported by
    `instr_truncated()`, so it surfaces as an operator warning instead of quietly
    shipping incomplete legal wording.
    """
    def build(text: str) -> str:
        return f'<updated_addr>{text}<a href="{url}">{url}</a></updated_addr>'

    if len(build(mandated + cta)) <= limit:
        return build(mandated + cta)
    if len(build(mandated)) <= limit:
        return build(mandated)
    t = mandated
    while t and len(build(t[:-1] + "…")) > limit:
        t = t[:-1]
    return build((t[:-1] + "…") if t else "…")


def instr_truncated(text: str) -> bool:
    """True when _fit_instr had to ellipsis-trim to fit the 500-char cap.

    The forward RDO wording is compliance text (guide §2.4) and currently lands at exactly
    500/500 with a 32-char token — zero spare. A longer PUBLIC_BASE_URL would silently eat
    the tail, so callers surface this as an operator-visible warning instead.
    """
    return "…" in text


def delivery_instructions(service_code: str, awb: dict, url: str) -> str:
    """Build the col-R delivery_instructions (fixed text + injected courier link, ≤500 chars)."""
    svc = CFG["services"][service_code]
    limit = CFG["link_char_limit"]
    if svc["direction"] == "return":
        return _fit_instr(CFG["rdo_text"]["return_delivery_short"], "", url, limit)
    return _fit_forward(CFG["rdo_text"]["forward"], CFG["rdo_text"].get("forward_cta", ""),
                        url, limit)


def _item_description(awb: dict) -> str:
    if awb["is_return"]:
        return CFG["fixed"]["item_description_return"]
    total = awb["collies"]
    parts = ", ".join(f'{p["po_number"]} ({p["koli"]})' for p in awb["po_lines"])
    return f"{parts} — {total} koli"


def _upload_row(service_code: str, awb: dict, trid: str, trids: list[str], today: str) -> dict:
    """One NV upload row = one AWB = one MPS order (guide §2.3: 1 WP = 1 MPS TRID = 1 SwipeAWB).

    `trid` is the bundle's tracking number (the SwipeAWB); `trids` are its child pieces,
    which live only in `requested_piece_tracking_numbers`.
    """
    svc = CFG["services"][service_code]
    wh = CFG["warehouse"]
    fx = CFG["fixed"]
    ts = fx["timeslot"]
    row = {
        "requested_tracking_number": trid,
        "global_shipper_id": CFG["master_shipper_id"],
        "service_type": fx["service_type"],
        "reference.merchant_order_number": awb["merchant_order_number"],
        "service_level": svc["service_level"],
        "to.address.country": "ID",
        "to.address.kecamatan": "",
        "to.address.province": "",
        "parcel_job.delivery_instructions": awb["delivery_instructions"],
        "parcel_job.delivery_start_date": today,
        "parcel_job.delivery_timeslot.start_time": ts["start_time"],
        "parcel_job.delivery_timeslot.end_time": ts["end_time"],
        "parcel_job.delivery_timeslot.timezone": ts["timezone"],
        "parcel_job.items.0.item_description": _item_description(awb),
        "parcel_job.items.0.is_dangerous_good": fx["is_dangerous_good"],
        "b2b.documents_required": fx["documents_required"],
        "bundle_information.total_quantity": str(awb["collies"]),
        "bundle_information.requested_piece_tracking_numbers": ", ".join(trids),
        "parcel_job.insured_value": fx["insured_value"],
        "corporate.branch_id": svc["branch_id"],
    }
    if awb["is_return"]:
        # Reverse job: from = pharmacy, to = SwipeRx WH.
        row.update({
            "from.name": awb["pharmacy_name"], "from.phone_number": awb["phone"],
            "from.address.address1": awb["address"], "from.address.country": "ID",
            "to.name": wh["name"], "to.phone_number": wh["phone"], "to.address.address1": wh["address1"],
            "to.address.kecamatan": wh["kecamatan"], "to.address.city": wh["city"],
            "to.address.province": wh["province"], "to.address.postcode": wh["postcode"],
            "parcel_job.dimensions.weight": "1", "parcel_job.is_pickup_required": fx["is_pickup_required_return"],
            "parcel_job.pickup_date": today, "parcel_job.pickup_timeslot.start_time": ts["start_time"],
            "parcel_job.pickup_timeslot.end_time": ts["end_time"], "parcel_job.pickup_timeslot.timezone": ts["timezone"],
            "parcel_job.pickup_instructions": CFG["rdo_text"]["return_pickup_instructions"],
        })
    else:
        # Forward: from = SwipeRx WH, to = pharmacy.
        row.update({
            "from.name": wh["name"], "from.phone_number": wh["phone"], "from.address.address1": wh["address1"],
            "from.address.country": wh["country"],
            "to.name": awb["pharmacy_name"], "to.phone_number": awb["phone"], "to.address.address1": awb["address"],
            "to.address.city": awb["city"], "to.address.postcode": awb["postcode"],
            "parcel_job.dimensions.weight": awb["weight"] or "1",
            "parcel_job.is_pickup_required": fx["is_pickup_required_forward"],
        })
    return row


def build_upload_xlsx(service_code: str, awbs: list[dict], today: str | None = None) -> bytes:
    """Build the Ninja Van upload workbook. Each AWB must already have delivery_instructions set."""
    today = today or date.today().isoformat()
    is_return = CFG["services"][service_code]["direction"] == "return"
    cols = FWD_COLS + RET_EXTRA_COLS if is_return else FWD_COLS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload"
    ws.append(cols)
    for awb in awbs:
        # ONE row per AWB (guide §2.3, rev. 27 Jul 2026). The previous build emitted a row
        # per collie, which created N separate orders each claiming to be a bundle of N.
        #
        # `children_mode="blank"` leaves AC empty so NV mints the children itself. That is
        # the C4 claim in OC_AWB_PARENT_CHECK §2, which has never actually been shipped —
        # every one of the 1035 children in the three accepted uploads was filled in. The
        # manual test module sets this per-AWB so one trial upload can settle it; the TMP
        # intake path never does, and stays on the shape with production evidence.
        trids = [] if awb.get("children_mode") == "blank" else piece_trids(awb)
        row = _upload_row(service_code, awb, awb["awb_id"], trids, today)
        ws.append([str(row.get(col, "")) for col in cols])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_links_csv(awbs: list[dict]) -> bytes:
    """AWB → courier link, plus the return item detail/invoice the courier app shows behind the link."""
    bio = io.StringIO()
    w = csv.writer(bio)
    w.writerow(["scope", "token", "url", "invoice", "item_detail"])
    for a in awbs:
        w.writerow([a["awb_id"], a.get("token", ""), a.get("url", ""), a.get("invoice", ""), a.get("item_detail", "")])
    return ("﻿" + bio.getvalue()).encode("utf-8")
