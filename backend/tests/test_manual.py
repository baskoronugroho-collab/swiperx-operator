"""Manual link builder (manual.py) — the phase-1 field-test instrument.

These pin the two things that make it useful as a test tool: what it hands you to paste
into Ninja, and the guarantee that it can never touch an AWB that came from a real intake.
"""


def _make(client, **over):
    body = {
        "service": "S1",
        "awb_id": "AWBMANUAL1",
        "pharmacy_name": "Apotek Uji Coba",
        "address": "Jl. Uji 1",
        "city": "Depok",
        "weight": "1.77",
        "collies": 3,
        "po_lines": [{"po_number": "PO-AAA", "koli": 2}, {"po_number": "PO-BBB", "koli": 1}],
    }
    body.update(over)
    return client.post("/api/manual/links", json=body)


def test_manual_link_needs_an_intake_role(client):
    assert client.post("/api/manual/links", json={"service": "S1", "awb_id": "X"}).status_code == 401
    client.post("/api/auth/dev-login", json={"email": "nobody@ninjavan.co"})
    assert _make(client).status_code == 403


def test_it_reproduces_the_real_column_contract(de_client):
    """The whole point of the tool: hand back exactly what goes into Ninja.

    Col A is the SwipeAWB; the children are PO-prefixed and numbered continuously across
    the AWB (see oc_engine.piece_trids — PO-prefixed is intended, not a bug).
    """
    r = _make(de_client)
    assert r.status_code == 201, r.text
    cols = r.json()["upload_columns"]
    assert cols["A_requested_tracking_number"] == "AWBMANUAL1"
    assert cols["D_merchant_order_number"] == "AWBMANUAL1"
    assert cols["AB_total_quantity"] == "3"
    assert cols["AC_piece_tracking_numbers"] == "PO-AAA-01, PO-AAA-02, PO-BBB-03"


def test_col_r_carries_the_link_and_stays_inside_the_budget(de_client):
    body = _make(de_client).json()
    assert body["url"] in body["delivery_instructions"]
    assert body["delivery_instructions"].endswith("</updated_addr>")
    assert body["instr_length"] <= body["instr_limit"]
    assert body["instr_truncated"] is False


def test_blank_children_mode_leaves_ac_empty(de_client):
    """C4 is still unproven, so the tool must be able to ship a blank AC on one AWB."""
    body = _make(de_client, children_mode="blank").json()
    assert body["children"] == []
    assert body["upload_columns"]["AC_piece_tracking_numbers"] == ""
    assert body["upload_columns"]["AB_total_quantity"] == "3"  # AB is still required


def test_the_generated_link_opens_the_courier_wizard(de_client):
    body = _make(de_client).json()
    r = de_client.get(f"/api/c/{body['token']}", follow_redirects=False)
    assert r.status_code == 307
    assert r.headers["location"] == f"/c/{body['token']}"


def test_duplicate_awb_is_rejected(de_client):
    assert _make(de_client).status_code == 201
    r = _make(de_client)
    assert r.status_code == 409
    assert r.json()["detail"] == "awb_already_exists"


def test_bad_input_is_rejected(de_client):
    assert _make(de_client, awb_id="").status_code == 400
    assert _make(de_client, awb_id="A" * 33).status_code == 400
    assert _make(de_client, service="S9").status_code == 400
    assert _make(de_client, children_mode="sideways").status_code == 400
    assert _make(de_client, delivery_date="31-12-2026").status_code == 400


def test_listing_and_deleting_only_ever_touch_test_awbs(de_client, awb):
    """`awb` fixture is an intake-style row (intake_id NULL is what marks a manual one).

    The fixture inserts without an intake_id, so it *would* be listed — the meaningful
    guarantee is that a row created by a real intake (intake_id set) is invisible here.
    """
    _make(de_client)
    ids = [row["awb_id"] for row in de_client.get("/api/manual/links").json()["links"]]
    assert "AWBMANUAL1" in ids

    assert de_client.delete("/api/manual/links/AWBMANUAL1").status_code == 204
    assert de_client.delete("/api/manual/links/AWBMANUAL1").status_code == 404
    # Deleting frees the identifier for another attempt — the reason delete exists.
    assert _make(de_client).status_code == 201


def test_an_intake_owned_awb_is_never_listed_or_deletable(de_client, dbs):
    import asyncio

    asyncio.get_event_loop_policy().new_event_loop().run_until_complete(
        dbs.execute(
            "INSERT INTO awb (awb_id, merchant_order_number, service_id, pharmacy_name, koli, "
            "link_token, status, is_return, intake_id) "
            "VALUES ('AWBFROMINTAKE','AWBFROMINTAKE','S1','Apotek Nyata',1,'tok_real','created',0,42)", (),
        )
    )
    ids = [row["awb_id"] for row in de_client.get("/api/manual/links").json()["links"]]
    assert "AWBFROMINTAKE" not in ids
    assert de_client.delete("/api/manual/links/AWBFROMINTAKE").status_code == 404


def test_one_row_workbook_is_downloadable(de_client):
    _make(de_client)
    r = de_client.get("/api/manual/links/AWBMANUAL1/upload.xlsx")
    assert r.status_code == 200
    assert "manual-AWBMANUAL1.xlsx" in r.headers["content-disposition"]

    import io

    import openpyxl

    ws = openpyxl.load_workbook(io.BytesIO(r.content)).worksheets[0]
    assert ws.max_row == 2, "exactly one data row"
    hdr = [c.value for c in ws[1]]
    row = {h: ws.cell(2, i + 1).value for i, h in enumerate(hdr)}
    assert row["requested_tracking_number"] == "AWBMANUAL1"
    assert row["bundle_information.requested_piece_tracking_numbers"] == (
        "PO-AAA-01, PO-AAA-02, PO-BBB-03"
    )
