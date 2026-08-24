"""Lane 3 — the reject-return pipeline (24 Aug rework).

    pending_validator -> pending_de_upload -> pending_print -> printed     (sebagian)
                      -> pending_de_upload -> rts_triggered                (semua)

These pin the two properties that make the pipeline trustworthy: nothing moves until the
Validator has looked at the photos, and each closing path refuses the other type's rows.
"""
import io

import openpyxl
import pytest

from conftest import photo


def _reject(client, awb, return_type, pcs=3):
    """Drive a real courier reject end to end, the way production creates the row."""
    t = awb["token"]
    dn = client.post(f"/api/c/{t}/capture", data={"doc_type": "delivery_note"}, files=photo()).json()
    client.patch(f"/api/c/{t}/capture/{dn['id']}", json={"signed_stamped": True})
    for doc in ("delivery_note", "rejected_goods", "awb_sticker"):
        client.post(f"/api/c/{t}/capture", data={"doc_type": doc}, files=photo())
    r = client.post(
        f"/api/c/{t}/submit",
        json={"outcome": "reject", "return_type": return_type, "reject_pcs": pcs},
    )
    assert r.status_code == 200, r.text
    return awb


@pytest.fixture()
def rejected(client, awb):
    return _reject(client, awb, "sebagian")


@pytest.fixture()
def fully_rejected(client, awb):
    return _reject(client, awb, "semua")


def _row(c):
    return c.get("/api/returns").json()["returns"][0]


def _validate(c, rid):
    """Validate as Vera, then hand the session back to Dewi.

    The fixtures share ONE TestClient (one cookie jar), so 'being the validator' is a
    login switch, not a separate client — exactly like one browser profile would be.
    """
    c.post("/api/auth/dev-login", json={"email": "vera.v@ninjavan.co"})
    r = c.post("/api/returns/validate", json={"ids": [rid]})
    c.post("/api/auth/dev-login", json={"email": "dewi.k@ninjavan.co"})
    assert r.status_code == 200, r.text
    return r.json()


# ------------------------------------------------------------ entry state ----
def test_worklist_requires_a_session(client, rejected):  # noqa: ARG001
    assert client.get("/api/returns").status_code == 401


def test_validator_role_can_see_the_worklist(validator_client, rejected):  # noqa: ARG001
    assert validator_client.get("/api/returns").status_code == 200


def test_row_starts_pending_validator_with_the_door_evidence(de_client, rejected):  # noqa: ARG001
    r = _row(de_client)
    assert r["stage"] == "pending_validator"
    assert r["reject_pcs"] == 3
    assert r["origin"] == "TMP_DEPOK" and r["origin_unknown"] is False
    kinds = {p["doc_type"] for p in r["proof_photos"]}
    assert {"delivery_note", "rejected_goods", "awb_sticker"} <= kinds


def test_reject_submit_requires_the_pcs_count(client, awb):
    t = awb["token"]
    dn = client.post(f"/api/c/{t}/capture", data={"doc_type": "delivery_note"}, files=photo()).json()
    client.patch(f"/api/c/{t}/capture/{dn['id']}", json={"signed_stamped": True})
    for doc in ("delivery_note", "rejected_goods", "awb_sticker"):
        client.post(f"/api/c/{t}/capture", data={"doc_type": doc}, files=photo())
    r = client.post(f"/api/c/{t}/submit", json={"outcome": "reject", "return_type": "sebagian"})
    assert r.status_code == 422
    assert r.json()["detail"] == "reject_pcs_required"


# -------------------------------------------------------------- validation ----
def test_de_cannot_validate(de_client, rejected):  # noqa: ARG001
    rid = _row(de_client)["id"]
    assert de_client.post("/api/returns/validate", json={"ids": [rid]}).status_code == 403


def test_validation_moves_the_row_to_pending_de_upload(de_client, rejected):  # noqa: ARG001
    rid = _row(de_client)["id"]
    assert _validate(de_client, rid)["updated"] == 1
    r = _row(de_client)
    assert r["stage"] == "pending_de_upload"
    assert r["validated_by_email"] == "vera.v@ninjavan.co"
    # Idempotent: re-validating an already-moved row is a no-op, not an error.
    assert _validate(de_client, rid)["updated"] == 0


def test_nothing_downstream_works_before_validation(de_client, rejected):  # noqa: ARG001
    rid = _row(de_client)["id"]
    assert de_client.get("/api/returns/export-oc.xlsx").status_code == 404
    assert de_client.post("/api/returns/mark-uploaded", json={"ids": [rid]}).json()["updated"] == 0
    assert de_client.post("/api/returns/rts", json={"ids": [rid]}).json()["updated"] == 0


# --------------------------------------------------- sebagian: OC pipeline ----
def test_partial_walks_export_upload_print(de_client, rejected):
    rid = _row(de_client)["id"]
    _validate(de_client, rid)

    # Export: one row, <SwipeAWB>-R01, addressed to the origin warehouse, pcs in col Y.
    r = de_client.get("/api/returns/export-oc.xlsx")
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.content)).worksheets[0]
    hdr = [c.value for c in ws[1]]
    row = {h: ws.cell(2, i + 1).value for i, h in enumerate(hdr)}
    assert ws.max_row == 2
    assert row["requested_tracking_number"] == f"{rejected['awb_id']}-R01"
    assert row["reference.merchant_order_number"] == rejected["awb_id"]
    assert row["to.address.city"] == "Depok"
    assert row["parcel_job.items.0.item_description"] == "3"

    # Mark uploaded -> Pending Print, with the -R01 stamped for IC's OPV2 search.
    assert de_client.post("/api/returns/mark-uploaded", json={"ids": [rid]}).json()["updated"] == 1
    r2 = _row(de_client)
    assert r2["stage"] == "pending_print"
    assert r2["return_awb_id"] == f"{rejected['awb_id']}-R01"
    # Once uploaded it leaves the export file.
    assert de_client.get("/api/returns/export-oc.xlsx").status_code == 404

    # Printed & labelled -> closed.
    assert de_client.post("/api/returns/mark-printed", json={"ids": [rid]}).json()["updated"] == 1
    assert _row(de_client)["stage"] == "printed"


async def _strip(dbs, awb_id):
    await dbs.execute("UPDATE awb SET origin = NULL WHERE awb_id = ?", (awb_id,))
    await dbs.execute("UPDATE return_parcel SET origin = NULL WHERE original_awb_id = ?", (awb_id,))


def test_origin_unknown_blocks_the_oc_export_until_bulk_set(de_client, dbs, rejected):  # noqa: ARG001
    import asyncio

    rid = _row(de_client)["id"]
    _validate(de_client, rid)
    # Strip the origin from both the row and its forward AWB — the pre-tracking case.
    asyncio.get_event_loop_policy().new_event_loop().run_until_complete(
        _strip(dbs, rejected["awb_id"])
    )
    r = _row(de_client)
    assert r["origin_unknown"] is True
    assert de_client.get("/api/returns/export-oc.xlsx").status_code == 404
    assert de_client.post("/api/returns/mark-uploaded", json={"ids": [rid]}).status_code == 409

    assert de_client.post(
        "/api/returns/origin", json={"ids": [rid], "origin": "TMP_SURABAYA"}
    ).json()["updated"] == 1
    assert _row(de_client)["origin_unknown"] is False
    assert de_client.get("/api/returns/export-oc.xlsx").status_code == 200


# ------------------------------------------------------- semua: RTS pipeline --
def test_full_refusal_closes_by_rts_and_never_prints(de_client, fully_rejected):
    rid = _row(de_client)["id"]
    assert _row(de_client)["closes_by"] == "rts"
    _validate(de_client, rid)

    # The full refusal never appears in the OC export — no new AWB exists for it.
    assert de_client.get("/api/returns/export-oc.xlsx").status_code == 404
    # And the print path refuses it outright.
    assert de_client.post("/api/returns/mark-uploaded", json={"ids": [rid]}).json()["updated"] == 0

    csv_text = de_client.get("/api/returns/export-rts.csv").text
    assert fully_rejected["awb_id"] in csv_text and ",no" in csv_text

    assert de_client.post("/api/returns/rts", json={"ids": [rid]}).json()["updated"] == 1
    r = _row(de_client)
    assert r["stage"] == "rts_triggered"
    assert r["return_awb_id"] is None  # the whole point: no second tracking number
    # Still exportable, now flagged as marked — the list matches what was just done.
    assert ",yes" in de_client.get("/api/returns/export-rts.csv").text
    # Marking twice is a no-op.
    assert de_client.post("/api/returns/rts", json={"ids": [rid]}).json()["updated"] == 0


def test_partial_rejects_are_invisible_to_the_rts_path(de_client, rejected):  # noqa: ARG001
    rid = _row(de_client)["id"]
    _validate(de_client, rid)
    assert de_client.post("/api/returns/rts", json={"ids": [rid]}).json()["updated"] == 0
    assert de_client.get("/api/returns/export-rts.csv").status_code == 404


# ------------------------------------------------------------------ misc ------
def test_stage_filter_and_bad_stage(de_client, rejected):  # noqa: ARG001
    rows = de_client.get("/api/returns?stage=pending_validator").json()["returns"]
    assert len(rows) == 1
    assert de_client.get("/api/returns?stage=pending_ack").status_code == 400


def test_csv_export_carries_the_full_trail(de_client, rejected):
    rid = _row(de_client)["id"]
    _validate(de_client, rid)
    de_client.post("/api/returns/mark-uploaded", json={"ids": [rid]})
    text = de_client.get("/api/returns/export.csv").text
    assert "validated_by" in text and "vera.v@ninjavan.co" in text
    assert f"{rejected['awb_id']}-R01" in text
    assert "pending_print" in text
