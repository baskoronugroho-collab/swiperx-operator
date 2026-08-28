"""OC engine — the MPS model from OC_TEMPLATE_AND_ENGINE_GUIDE.md §2.3 (rev. 27 Jul 2026).

These pin the rules that broke the DE team's spreadsheet converter, so a future change to
the tracking-number scheme has to break a test rather than a live upload.
"""
import io

import openpyxl
import pytest

import oc_engine as e


def _awb(**kw):
    base = {
        "awb_id": "AWB02U24V", "merchant_order_number": "AWB02U24V", "is_return": False,
        "pharmacy_name": "Apotek Uji", "address": "Jl. Uji 1", "city": "Depok",
        "postcode": "16451", "phone": "0812", "weight": "4.5", "collies": 4,
        "invoice": None, "item_detail": None, "delivery_instructions": "x",
        "po_lines": [{"po_number": "PO1", "koli": 2}, {"po_number": "PO2", "koli": 1},
                     {"po_number": "PO3", "koli": 1}],
    }
    base.update(kw)
    return base


def test_children_are_po_prefixed_and_numbered_continuously():
    """The guide's worked example, verbatim: numbering runs across the AWB, not per PO.

    PO-prefixed is intended — NV allows custom piece ids and the children still attach to
    their parent via the manual template (confirmed 10 and 18 Aug 2026). A change to
    parent-prefixed children was made and reverted on 10 Aug; this test is what should stop
    it happening a third time.
    """
    assert e.piece_trids(_awb()) == ["PO1-01", "PO1-02", "PO2-03", "PO3-04"]


def test_numbering_runs_across_the_awb_not_per_po():
    """A multi-koli PO shifts every ordinal after it — the counter belongs to the AWB."""
    awb = _awb(collies=5, po_lines=[{"po_number": "PO1", "koli": 3},
                                    {"po_number": "PO2", "koli": 2}])
    assert e.piece_trids(awb) == ["PO1-01", "PO1-02", "PO1-03", "PO2-04", "PO2-05"]


def test_children_are_zero_padded_past_nine():
    awb = _awb(collies=12, po_lines=[{"po_number": "POA", "koli": 12}])
    trids = e.piece_trids(awb)
    assert trids[0] == "POA-01"
    assert trids[8] == "POA-09"
    assert trids[9] == "POA-10"  # zero-padded to 2, never "-9"/"-10" mixed widths
    assert len(trids) == 12


def test_single_koli_awb_still_gets_a_numbered_child():
    awb = _awb(collies=1, po_lines=[{"po_number": "POX", "koli": 1}])
    assert e.piece_trids(awb) == ["POX-01"]


def test_return_awb_gets_one_piece_off_the_presupplied_awb():
    awb = _awb(awb_id="AWBR-10909-2026-5-16", is_return=True, collies=1,
               po_lines=[{"po_number": "POR", "koli": 1}])
    assert e.piece_trids(awb) == ["AWBR-10909-2026-5-16-01"]


def _rows(service, awbs):
    wb = openpyxl.load_workbook(io.BytesIO(e.build_upload_xlsx(service, awbs, "2026-08-01")))
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    return header, [dict(zip(header, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]


def test_upload_emits_exactly_one_row_per_awb():
    """1 WP = 1 MPS TRID = 1 SwipeAWB. The pre-27-Jul build emitted a row per collie,
    which would create N separate orders each claiming to be a bundle of N."""
    _, rows = _rows("S1", [_awb(), _awb(awb_id="AWB02U51N", merchant_order_number="AWB02U51N",
                                 collies=1, po_lines=[{"po_number": "POZ", "koli": 1}])])
    assert len(rows) == 2


def test_tracking_number_equals_the_swipe_awb_and_matches_merchant_order_number():
    _, rows = _rows("S1", [_awb()])
    r = rows[0]
    assert r["requested_tracking_number"] == "AWB02U24V"
    assert r["reference.merchant_order_number"] == "AWB02U24V"


def test_bundle_quantity_and_pieces_agree():
    _, rows = _rows("S1", [_awb()])
    r = rows[0]
    assert r["bundle_information.total_quantity"] == "4"
    pieces = r["bundle_information.requested_piece_tracking_numbers"].split(", ")
    assert pieces == ["PO1-01", "PO1-02", "PO2-03", "PO3-04"]
    assert len(pieces) == int(r["bundle_information.total_quantity"])
    # Col A stays the SwipeAWB even though the children carry PO prefixes.
    assert r["requested_tracking_number"] == "AWB02U24V"


def test_weight_is_the_awb_total_not_a_hardcoded_one():
    """The sample templates hardcode W=1; the guide says read the real AWB weight."""
    _, rows = _rows("S1", [_awb(weight="16.1")])
    assert rows[0]["parcel_job.dimensions.weight"] == "16.1"


def test_item_description_carries_the_po_cross_check():
    _, rows = _rows("S1", [_awb()])
    assert rows[0]["parcel_job.items.0.item_description"] == "PO1 (2), PO2 (1), PO3 (1) — 4 koli"


@pytest.mark.parametrize(
    ("service", "level", "branch"), [("S1", "STANDARD", "1"), ("S2", "SAMEDAY", "3")]
)
def test_s1_and_s2_differ_only_by_service_level_and_branch(service, level, branch):
    _, rows = _rows(service, [_awb()])
    r = rows[0]
    assert r["service_level"] == level
    assert r["corporate.branch_id"] == branch
    # Col B is ALWAYS the master; the service is selected by the branch id, never here.
    assert r["global_shipper_id"] == "11398423"


def test_s1_is_named_regular_b2br():
    s1 = next(s for s in e.services() if s["code"] == "S1")
    assert s1["name"] == "Regular B2BR"


PROD_URL = "https://swiperx-operator.ninjavan.apps.substrait.build/c/" + "x" * 32


def test_forward_rdo_text_fits_the_500_char_column_without_truncation():
    """Col R is compliance text with a hard cap. Against the real production host the
    mandated wording must survive intact."""
    out = e.delivery_instructions("S1", _awb(), PROD_URL)
    label = e.CFG["rdo_text"]["forward_link_label"]
    assert len(out) <= e.CFG["link_char_limit"]
    assert not e.instr_truncated(out), "mandated RDO wording was trimmed"
    # The FINAL template (Baskoro, 26 Aug 2026): mandated wording first as plain text
    # OUTSIDE the wrapper, then the whole retur sentence + visible URL as ONE anchor —
    # the entire call-to-action is the tap target, not a bare URL.
    assert out == (f'{e.CFG["rdo_text"]["forward"]} '
                   f'<updated_addr><a href="{PROD_URL}">{label}{PROD_URL}</a></updated_addr>')


def test_an_oversized_url_trims_loudly_never_silently():
    """If a host ever grows past the budget, the mandated wording is ellipsised AND
    flagged by instr_truncated(), so it lands as an operator warning — the field must
    never quietly ship incomplete legal wording. (The old drop-the-CTA middle ground died
    with the 24 Aug link-first change: the CTA pointed at a trailing link that no longer
    exists, and retiring it bought ~10 chars of spare.)"""
    limit = e.CFG["link_char_limit"]
    label = e.CFG["rdo_text"]["forward_link_label"]
    # Size the URL (counted twice) so ~100 chars remain for the mandated wording — enough
    # that the ellipsised text still fits, too little for the full block.
    overhead = len(' <updated_addr><a href=""></a></updated_addr>') + len(label)
    url_len = (limit - overhead - 100) // 2
    url = "https://" + "h" * (url_len - len("https:///c/") - 32) + "/c/" + "x" * 32
    out = e.delivery_instructions("S1", _awb(), url)
    assert len(out) <= limit
    assert e.instr_truncated(out), "an over-budget field must be flagged, not silent"
    # The link sentence is never the thing trimmed — it is the point of the field.
    assert out.endswith(f'<a href="{url}">{label}{url}</a></updated_addr>')


def test_both_origins_carry_confirmed_real_contact_details():
    """The `to.*` block of every return OC comes straight from these values — no
    normalisation happens anywhere between config and the Ninja upload column.

    History: TMP Surabaya shipped from 19 Aug 2026 with Depok's phone number as a
    placeholder, flagged `phone_confirmed: false`, because nobody had the real one. It
    reached col K of every Surabaya return OC. Baskoro supplied both sites' real details on
    28 Aug 2026 and they are pinned here so a placeholder can never be mistaken for a
    confirmed value again.

    The two phone FORMATS differ on purpose and are his call, not a typo to tidy: Depok is
    the local 0-prefixed form, Surabaya the 62 country-code form. If Ninja ever rejects one,
    fix it here and in oc_config.json together.
    """
    depok = e.CFG["origins"]["TMP_DEPOK"]
    surabaya = e.CFG["origins"]["TMP_SURABAYA"]

    # One shipper account name across both sites; the SITE is identified by the address.
    account = "PT Teknologi Medika Pratama (SwipeRx) (B2BR)"
    assert depok["name"] == account
    assert surabaya["name"] == account

    assert depok["phone"] == "087733785699"
    assert surabaya["phone"] == "6285143002578"
    assert depok["address1"].startswith("KUBIK LOGISTICS - ")
    assert surabaya["address1"].startswith("PT Teknologi Medika Pratama - Surabaya, ")

    # No origin may ship to production carrying a placeholder contact number.
    for code, o in e.CFG["origins"].items():
        assert o.get("phone_confirmed") is True, f"{code} still has an unconfirmed phone"
