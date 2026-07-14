"""Engine-only local test — runs backend/oc_engine.py against the real TMP files with
NO database, NO server. Exercises the exact M1 transform (parse -> counts -> upload.xlsx
-> links.csv) and writes per-service CSV output for Baskoro to open in Excel.

Run:  C:/Users/NXP/anaconda3/python.exe tools/oc_local_test.py
Output: tools/local-test-output/S1/upload.csv + links.csv (and S2, S3)
"""
import csv
import io
import os
import secrets
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND = os.path.join(ROOT, "backend")
OUT_DIR = os.path.join(ROOT, "tools", "local-test-output")
sys.path.insert(0, BACKEND)
os.chdir(BACKEND)                       # so oc_engine finds oc_config.json (relative)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import oc_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci

T = os.path.join(ROOT, "OC Template")
CASES = {
    "S1": ("[Template Swipe Fwd Reg] TMP Batch Order Ninja Express Intercity (FULL) 10-06-26.xlsx", 57, 144),
    "S2": ("[Template Swipe FWD Sameday] TMP Batch Ninja Depok 30-06-26 .xlsx", 79, 210),
    "S3": ("[Template Swipe PU Return] TMP Batch Special Case Ninja 15-06-26.xlsx", 23, 23),
}

BASE = "http://localhost:8000"   # fake link base for the local test (no server/DB involved)


def xlsx_to_csv_rows(xlsx_bytes: bytes) -> list[list[str]]:
    ws = load_workbook(io.BytesIO(xlsx_bytes)).active
    return [[c.value if c.value is not None else "" for c in row] for row in ws.iter_rows()]


def write_csv(path: str, rows: list[list[str]]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)


os.makedirs(OUT_DIR, exist_ok=True)
ok = True
for svc, (fname, exp_awb, exp_pieces) in CASES.items():
    data = open(os.path.join(T, fname), "rb").read()
    res = oc_engine.parse(data, svc)
    awbs = res["awbs"]
    pieces = sum(len(oc_engine.piece_trids(a)) for a in awbs)
    for a in awbs:
        a["token"] = secrets.token_urlsafe(24)
        a["url"] = f"{BASE}/api/c/{a['token']}"
        a["delivery_instructions"] = oc_engine.delivery_instructions(svc, a, a["url"])

    xlsx = oc_engine.build_upload_xlsx(svc, awbs)
    links_csv_bytes = oc_engine.build_links_csv(awbs)

    svc_dir = os.path.join(OUT_DIR, svc)
    os.makedirs(svc_dir, exist_ok=True)
    write_csv(os.path.join(svc_dir, "upload.csv"), xlsx_to_csv_rows(xlsx))
    with open(os.path.join(svc_dir, "links.csv"), "wb") as f:
        f.write(links_csv_bytes)

    ws = load_workbook(io.BytesIO(xlsx)).active
    maxR = max((len(ws.cell(r, ci("R")).value or "") for r in range(2, ws.max_row + 1)), default=0)
    tag = "OK " if (len(awbs) == exp_awb and pieces == exp_pieces and res["errors"] == []) else "!! "
    ok &= tag == "OK "
    print(f"{tag}{svc}: {len(awbs)} AWBs / {pieces} pieces (expected {exp_awb}/{exp_pieces}), "
          f"errors={len(res['errors'])}, rows={ws.max_row}, B2={ws.cell(2,ci('B')).value}, "
          f"AE2={ws.cell(2,ci('AE')).value}, maxR={maxR}  -> {svc_dir}\\upload.csv + links.csv")

print(f"\nOutput written to: {OUT_DIR}")
print("ALL GREEN" if ok else "MISMATCH — see !! rows")
sys.exit(0 if ok else 1)
